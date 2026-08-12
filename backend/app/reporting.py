"""Phase 5 — daily audit report computation and Excel export (§3.4).

`compute_report` derives the §3.4 metrics from the day's data:
- Stops reviewed = stops with a final audit decision (passed/failed) whose
  `audited_at` falls on the report date.
- Routes audited = distinct routes among those stops.
- Issues = SOP violations whose `created_at` falls on the report date,
  broken down by severity (Critical/Major/Minor) and source
  (Dispatch/Driver errors) per §3.2 and §3.7.

`build_xlsx` renders the report exactly in the §3.4 layout, plus a
Violations Detail sheet listing the day's issues.

`build_pickup_sheet_xlsx` exports the day's pickup sheets for ALL drivers
organized by state: each state name is a section heading, beneath it one row
per individual record (driver name, state, clinic ID, fare, status, audit
outcome) with one column per SOP checklist item so each record shows how it
was judged (Pass / Fail / N/A). Stop rows are color-coded by audit outcome:
green when the audit passed, orange when it carries a minor flag, red when it
carries a major/critical flag.
"""
from collections import Counter, OrderedDict
from datetime import date, datetime, time
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from . import models

HEADER_FILL = PatternFill("solid", fgColor="1E293B")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)

# Pickup-sheet export fills — the §3.4 color key.
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")  # green — audit passed
MINOR_FILL = PatternFill("solid", fgColor="FFE0B2")  # orange — minor flag
MAJOR_FILL = PatternFill("solid", fgColor="FFC7CE")  # red — major/critical flag

# Per-checklist-item judgment colors (the SOP columns of the export).
SOP_PASS_FILL = PatternFill("solid", fgColor="E8F5E9")  # pale green
SOP_PASS_FONT = Font(bold=True, size=10, color="1B7A3D")
SOP_FAIL_FILL = PatternFill("solid", fgColor="FFEBEE")  # pale red
SOP_FAIL_FONT = Font(bold=True, size=10, color="C62828")
SOP_NA_FILL = PatternFill("solid", fgColor="F1F5F9")  # pale slate
SOP_NA_FONT = Font(size=10, color="64748B")
SOP_EMPTY_FONT = Font(size=10, color="94A3B8")

THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)
STATE_TOTAL_FILL = PatternFill("solid", fgColor="F8FAFC")
STATE_TOTAL_FONT = Font(italic=True, size=9, color="475569")

# Canonical SOP order (mirrors routers/checklist.py — the source of truth).
# Every item becomes a column so the sheet reflects ALL SOPs of the pickup
# sheet; stops that haven't been audited yet show "—".
SOP_TEMPLATE_ORDER = [
    "Pickup Photo",
    "Delivery Photo",
    "Building Photo",
    "Lockbox Photo",
    "Proof Slip",
    "Shipping Label Visible",
    "Barcode Readable",
    "Package Count Verified",
    "Correct Clinic",
    "Correct Destination",
    "Timestamp Verified",
    "Dispatch Verified",
    "Driver Notes Reviewed",
    "SOP Followed",
    "Inside Lockbox Photo",
    "Outside Lockbox Photo",
    "Front of Clinic Photo",
    "Specimen Count Photo",
    "Reception Area Photo",
    "Package Photo",
    "Clinic Entrance Photo",
]

_SOP_STATUS_LABEL = {
    models.CHK_PASS: "Pass",
    models.CHK_FAIL: "Fail",
    models.CHK_NA: "N/A",
}


def _day_bounds(d: date) -> tuple[datetime, datetime]:
    return datetime.combine(d, time.min), datetime.combine(d, time.max)


def compute_report(db, report_date: date) -> dict:
    """Derive the §3.4 metrics for one date (no writes)."""
    start, end = _day_bounds(report_date)

    reviewed = (
        db.query(models.Stop)
        .filter(
            models.Stop.audit_status.in_([models.AUDIT_PASSED, models.AUDIT_FAILED]),
            models.Stop.audited_at >= start,
            models.Stop.audited_at <= end,
        )
        .all()
    )
    passed = sum(1 for s in reviewed if s.audit_status == models.AUDIT_PASSED)
    failed = len(reviewed) - passed

    sev = {"critical": 0, "major": 0, "minor": 0}
    src = {"driver": 0, "dispatch": 0}
    for v in (
        db.query(models.Violation)
        .filter(
            models.Violation.created_at >= start,
            models.Violation.created_at <= end,
        )
        .all()
    ):
        sev[v.severity] = sev.get(v.severity, 0) + 1
        src[v.source] = src.get(v.source, 0) + 1

    return {
        "routes_audited": len({s.route_id for s in reviewed}),
        "stops_reviewed": len(reviewed),
        "passed": passed,
        "failed": failed,
        "critical": sev["critical"],
        "major": sev["major"],
        "minor": sev["minor"],
        "dispatch_errors": src["dispatch"],
        "driver_errors": src["driver"],
    }


def violations_for_date(db, report_date: date):
    """The day's violations, newest first (for the detail sheet / report page)."""
    start, end = _day_bounds(report_date)
    return (
        db.query(models.Violation)
        .filter(
            models.Violation.created_at >= start,
            models.Violation.created_at <= end,
        )
        .order_by(models.Violation.created_at.desc())
        .all()
    )


def _header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center")


def build_xlsx(report: dict, violations: list[dict]) -> BytesIO:
    """Render the report in the §3.4 layout + a violations detail sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Audit Report"

    ws["A1"] = "AGL QA — Daily Audit Report"
    ws["A1"].font = TITLE_FONT
    auditor = report.get("auditor_name") or "—"
    ws["A2"] = f"Date: {report['report_date']}    Auditor: {auditor}"

    # §3.4 — summary row 1
    headers1 = [
        "Date",
        "Auditor",
        "Routes Audited",
        "Stops Reviewed",
        "Total Passed",
        "Total Failed",
    ]
    row1 = [
        report["report_date"].isoformat(),
        auditor,
        report["routes_audited"],
        report["stops_reviewed"],
        report["passed"],
        report["failed"],
    ]
    for i, (h, v) in enumerate(zip(headers1, row1), start=1):
        _header(ws.cell(row=4, column=i, value=h))
        ws.cell(row=5, column=i, value=v)

    # §3.4 — issues row 2
    headers2 = [
        "Critical Issues",
        "Major Issues",
        "Minor Issues",
        "Dispatch Errors",
        "Driver Errors",
    ]
    row2 = [
        report["critical"],
        report["major"],
        report["minor"],
        report["dispatch_errors"],
        report["driver_errors"],
    ]
    for i, (h, v) in enumerate(zip(headers2, row2), start=1):
        _header(ws.cell(row=7, column=i, value=h))
        ws.cell(row=8, column=i, value=v)

    # Recommendations
    ws["A10"] = "Recommendations"
    ws["A10"].font = Font(bold=True)
    ws.merge_cells("A11:F14")
    rec = ws["A11"]
    rec.value = report.get("recommendations") or ""
    rec.alignment = Alignment(wrap_text=True, vertical="top")

    for i, w in enumerate([12, 16, 15, 14, 13, 13], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Violations detail sheet
    ws2 = wb.create_sheet("Violations Detail")
    headers = [
        "ID",
        "Logged",
        "Severity",
        "Source",
        "Category",
        "Description",
        "Route",
        "Stop",
        "Escalated",
        "Resolved",
    ]
    for i, h in enumerate(headers, start=1):
        _header(ws2.cell(row=1, column=i, value=h))

    def _stamp(v: datetime | None) -> str:
        return v.strftime("%Y-%m-%d %H:%M") if v else ""

    for r, v in enumerate(violations, start=2):
        ws2.cell(row=r, column=1, value=v.get("id"))
        ws2.cell(row=r, column=2, value=_stamp(v.get("created_at")))
        ws2.cell(row=r, column=3, value=(v.get("severity") or "").title())
        ws2.cell(row=r, column=4, value=(v.get("source") or "").title())
        ws2.cell(row=r, column=5, value=v.get("category"))
        ws2.cell(row=r, column=6, value=v.get("description"))
        ws2.cell(
            row=r,
            column=7,
            value=f"#{v['route_id']}" if v.get("route_id") else "",
        )
        ws2.cell(row=r, column=8, value=v.get("stop_label") or "")
        ws2.cell(row=r, column=9, value=_stamp(v.get("escalated_at")))
        ws2.cell(row=r, column=10, value=_stamp(v.get("resolved_at")))
    ws2.freeze_panes = "A2"
    for i, w in enumerate([6, 16, 10, 10, 24, 60, 8, 28, 18, 18], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Pickup sheet Excel export — all drivers for a date, organized by state
# ---------------------------------------------------------------------------


def _audit_label(status: str) -> str:
    return (status or "").replace("_", " ").title()


def _stop_row_fill(stop: models.Stop, violations: list) -> PatternFill | None:
    """Color key for one driver/clinic row: red (major/critical flag) > orange
    (minor flag) > green (audit passed). Unaudited, flag-free rows stay plain."""
    severities = [v.severity for v in violations]
    if any(s in (models.SEV_MAJOR, models.SEV_CRITICAL) for s in severities):
        return MAJOR_FILL
    if models.SEV_MINOR in severities:
        return MINOR_FILL
    if stop.audit_status == models.AUDIT_PASSED:
        return PASS_FILL
    return None


def _violations_summary(violations: list) -> str:
    """e.g. "3 (2 major, 1 minor)" — critical counts toward the major tier."""
    if not violations:
        return ""
    counts = Counter(v.severity for v in violations)
    parts = []
    for sev in (models.SEV_CRITICAL, models.SEV_MAJOR, models.SEV_MINOR):
        if counts.get(sev):
            parts.append(f"{counts[sev]} {sev}")
    return f"{len(violations)} ({', '.join(parts)})"


def build_pickup_sheet_xlsx(db, report_date: date) -> BytesIO:
    """One workbook for the date: each state is a section heading; beneath it
    every stop carries the driver name, state, clinic ID, fare and the rest of
    the pickup-sheet columns, then one column per SOP (checklist item) so each
    audited record shows exactly how it was judged.

    Formatting: record columns carry the audit color key (green = passed,
    orange = minor flag, red = major/critical flag), each checklist cell is
    colored by its own result (Pass green / Fail red / N/A gray), borders and
    a per-state total row keep it print-ready, and the header stays frozen.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Pickup Sheet"
    ws.sheet_properties.tabColor = "1E293B"

    routes = (
        db.query(models.Route)
        .filter(models.Route.route_date == report_date)
        .order_by(models.Route.route_date, models.Route.id)
        .all()
    )
    stops = [s for r in routes for s in sorted(r.stops, key=lambda s: s.sequence)]

    violations_by_stop: dict[int, list] = {}
    checklist_by_stop: dict[int, dict[str, str]] = {}
    if stops:
        stop_ids = [s.id for s in stops]
        for v in (
            db.query(models.Violation)
            .filter(models.Violation.stop_id.in_(stop_ids))
            .all()
        ):
            violations_by_stop.setdefault(v.stop_id, []).append(v)
        for item in (
            db.query(models.ChecklistItem)
            .filter(models.ChecklistItem.stop_id.in_(stop_ids))
            .all()
        ):
            checklist_by_stop.setdefault(item.stop_id, {})[item.item_name] = item.status

    # Column layout — one row per individual record: driver name, state,
    # clinic ID, fare, then every audit checklist item as a judging column.
    columns: list[tuple[str, int]] = [
        ("Driver", 18),
        ("State", 10),
        ("Stop", 6),
        ("Type", 10),
        ("Clinic ID", 10),
        ("Pickup location", 16),
        ("FedEx cutoff", 12),
        ("Fare ($)", 10),
        ("Status", 12),
        ("Audit", 12),
        ("Violations", 16),
    ]
    columns += [(name, 16) for name in SOP_TEMPLATE_ORDER]
    columns.append(("Notes", 30))
    header_labels = [c[0] for c in columns]
    # Base record columns (Driver … Violations) — everything before the
    # SOP judging columns. Derived so the layout can't drift from the list.
    base_count = len(columns) - len(SOP_TEMPLATE_ORDER) - 1
    sop_end = base_count + len(SOP_TEMPLATE_ORDER)
    fare_col = header_labels.index("Fare ($)") + 1

    ws["A1"] = "AGL QA — Pickup Sheet (all drivers)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"Date: {report_date.isoformat()}    "
        f"{len(routes)} route(s) · {len(stops)} stop(s)"
    )
    ws["A2"].font = Font(size=10, color="475569")

    # Color legend
    legend_row = 4
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True, size=9)
    for col, (label, fill) in enumerate(
        [
            ("Audit passed", PASS_FILL),
            ("Minor flag", MINOR_FILL),
            ("Major/Critical flag", MAJOR_FILL),
        ],
        start=2,
    ):
        cell = ws.cell(row=legend_row, column=col, value=label)
        cell.fill = fill
        cell.font = Font(bold=True, size=9)

    # Group stops by state, then by driver within each state.
    groups = OrderedDict()
    for s in stops:
        state = s.route.state if s.route and s.route.state else None
        key = (state.id if state else None, state.name if state else None, state.code if state else None)
        groups.setdefault(key, []).append(s)
    sorted_groups = sorted(groups.items(), key=lambda kv: (kv[0][1] is None, kv[0][1] or ""))

    row = legend_row + 2
    for (state_id, state_name, state_code), group_stops in sorted_groups:
        label = state_name or "No state assigned"
        if state_code:
            label = f"{state_name} ({state_code})"

        # State name heading — merged across all columns. Fill every cell in
        # the range BEFORE merging: openpyxl turns the non-anchor cells of a
        # merged range into read-only MergedCell objects, so styling them after
        # merge_cells() raises AttributeError. The anchor keeps its fill and
        # Excel renders the whole band in it.
        for col_i in range(1, len(columns) + 1):
            ws.cell(row=row, column=col_i).fill = HEADER_FILL
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=len(columns)
        )
        cell = ws.cell(row=row, column=1, value=f"{label} — {len(group_stops)} stop(s)")
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        row += 1

        # Column headers
        for col_i, name in enumerate(header_labels, start=1):
            _header(ws.cell(row=row, column=col_i, value=name))
        row += 1

        group_stops.sort(
            key=lambda s: (
                (s.route.driver.name if s.route and s.route.driver else ""),
                s.sequence,
            )
        )
        prev_driver_id = None
        for s in group_stops:
            route = s.route
            driver = route.driver if route else None
            row_state = route.state.code if route and route.state else ""
            violations = violations_by_stop.get(s.id, [])
            fill = _stop_row_fill(s, violations)
            checklist = checklist_by_stop.get(s.id, {})
            is_new_driver = (driver.id if driver else None) != prev_driver_id
            prev_driver_id = driver.id if driver else None

            clinic_label = ""
            if s.location_type == models.LOC_CLINIC:
                clinic_label = str(s.clinic_id) if s.clinic_id is not None else (s.clinic_ref or "")
            elif s.carrier_id is not None:
                clinic_label = str(s.carrier_id)

            values = [
                driver.name if driver else "",
                row_state,
                s.sequence,
                "Pickup" if s.stop_type == models.STOP_PICKUP else "Delivery",
                clinic_label,
                s.pickup_location or "",
                s.fedex_cutoff.isoformat()[:5] if s.fedex_cutoff else "",
                round(s.charge.driver_pay, 2) if s.charge and s.charge.driver_pay is not None else None,
                (s.status or "").title(),
                _audit_label(s.audit_status),
                _violations_summary(violations),
            ]
            values += [
                _SOP_STATUS_LABEL.get(checklist.get(name), "—") for name in SOP_TEMPLATE_ORDER
            ]
            values.append(s.notes or "")

            for col_i, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_i, value=value)
                cell.border = THIN_BORDER
                if col_i <= base_count:
                    # Record part — audit color key + light alignment polish.
                    if fill is not None:
                        cell.fill = fill
                    if col_i == fare_col and isinstance(value, (int, float)):
                        cell.number_format = "0.00"
                        cell.alignment = Alignment(horizontal="right")
                    # Left-aligned record columns: Driver(1), State(2),
                    # Pickup location(6), Violations(11) — the rest center.
                    elif col_i in (1, 2, 6, 11):
                        cell.alignment = Alignment(vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_i <= sop_end:
                    # Checklist judging columns — color each cell by its result.
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if value == "Pass":
                        cell.fill = SOP_PASS_FILL
                        cell.font = SOP_PASS_FONT
                    elif value == "Fail":
                        cell.fill = SOP_FAIL_FILL
                        cell.font = SOP_FAIL_FONT
                    elif value == "N/A":
                        cell.fill = SOP_NA_FILL
                        cell.font = SOP_NA_FONT
                    else:
                        cell.font = SOP_EMPTY_FONT
                else:
                    # Notes — wrapped, no fill so it stays readable.
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            if is_new_driver and driver:
                ws.cell(row=row, column=1).font = Font(bold=True, size=10)
            row += 1

        # Per-state total row (merged, subtle) — stops / passed / flagged / fare.
        # Fill BEFORE merge (merged non-anchor cells are read-only in openpyxl).
        passed_count = sum(1 for s in group_stops if s.audit_status == models.AUDIT_PASSED)
        flagged_count = sum(1 for s in group_stops if violations_by_stop.get(s.id))
        fare_total = sum(
            round(s.charge.driver_pay, 2) for s in group_stops if s.charge and s.charge.driver_pay
        )
        for col_i in range(1, len(columns) + 1):
            ws.cell(row=row, column=col_i).fill = STATE_TOTAL_FILL
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=len(columns)
        )
        total = ws.cell(
            row=row, column=1,
            value=(
                f"State total — {len(group_stops)} stop(s) · {passed_count} passed · "
                f"{flagged_count} flagged · ${fare_total:.2f} total fare"
            ),
        )
        total.font = STATE_TOTAL_FONT
        row += 1

    for i, (_, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Keep the title + legend visible while scrolling through the records.
    ws.freeze_panes = "B5"

    # Print-ready: landscape, fit to one page wide.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
